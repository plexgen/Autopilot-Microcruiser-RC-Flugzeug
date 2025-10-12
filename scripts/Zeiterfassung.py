#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Zeiterfassung (Tkinter) mit SQLite-DB, Excel/CSV-Export und interaktivem HTML-Report.
- Excel/CSV: Summen je Task + GESAMT-Zeile (Plan/Ist/Delta)
- HTML: kompakt skaliert (max-width 1200px, 420px Chart-Höhe) + Summen-Kacheln oben
"""
import os
import json
import html
import sqlite3
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk

DB_PATH = os.path.join(os.path.dirname(__file__), "zeiterfassung.db")
DATE_FMT = "%Y-%m-%d %H:%M:%S"
DATE_ONLY = "%Y-%m-%d"


# =========================
# DB Setup & Migration
# =========================
def now_str():
    return datetime.now().strftime(DATE_FMT)


def init_db():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL,
            planned_minutes INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS time_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            start_time TEXT,
            end_time TEXT,
            manual_minutes INTEGER NOT NULL DEFAULT 0,
            note TEXT,
            created_at TEXT,
            FOREIGN KEY(task_id) REFERENCES tasks(id)
        )
        """
    )
    con.commit()
    migrate_add_column(con, 'tasks', 'planned_minutes', "INTEGER NOT NULL DEFAULT 0")
    migrate_add_column(con, 'time_entries', 'created_at', "TEXT")
    cur.execute(
        "UPDATE time_entries "
        "SET created_at = COALESCE(created_at, start_time, end_time, ?) "
        "WHERE created_at IS NULL",
        (now_str(),)
    )
    con.commit()
    return con


def migrate_add_column(con, table, column, coldef):
    cur = con.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coldef}")
        con.commit()


# =========================
# DB-Helper
# =========================
def get_tasks(con):
    cur = con.cursor()
    cur.execute("SELECT id, name, created_at, planned_minutes FROM tasks ORDER BY name COLLATE NOCASE")
    return cur.fetchall()


def create_task(con, name, planned_hours=0, planned_minutes=0):
    if not name.strip():
        raise ValueError("Task-Name darf nicht leer sein.")
    pmin = max(0, int(planned_hours) * 60 + int(planned_minutes))
    cur = con.cursor()
    cur.execute(
        "INSERT INTO tasks(name, created_at, planned_minutes) VALUES (?, ?, ?)",
        (name.strip(), now_str(), pmin),
    )
    con.commit()


def get_active_entry(con):
    cur = con.cursor()
    cur.execute(
        "SELECT id, task_id, start_time FROM time_entries "
        "WHERE end_time IS NULL AND manual_minutes=0 LIMIT 1"
    )
    return cur.fetchone()


def start_task(con, task_id):
    active = get_active_entry(con)
    if active is not None:
        _, active_task_id, _ = active
        if active_task_id != task_id:
            raise RuntimeError("Ein anderer Task ist bereits aktiv.")
        return
    cur = con.cursor()
    cur.execute(
        "INSERT INTO time_entries(task_id, start_time, created_at) VALUES (?, ?, ?)",
        (task_id, now_str(), now_str()),
    )
    con.commit()


def stop_task(con, task_id):
    cur = con.cursor()
    cur.execute(
        "SELECT id FROM time_entries "
        "WHERE task_id=? AND end_time IS NULL AND manual_minutes=0 "
        "ORDER BY id DESC LIMIT 1",
        (task_id,),
    )
    row = cur.fetchone()
    if row is None:
        raise RuntimeError("Dieser Task ist nicht aktiv.")
    cur.execute("UPDATE time_entries SET end_time=? WHERE id=?", (now_str(), row[0]))
    con.commit()


def add_manual_time(con, task_id, hours, minutes, note="Manuell"):
    total_minutes = max(0, int(hours) * 60 + int(minutes))
    if total_minutes <= 0:
        raise ValueError("Bitte gültige Stunden/Minuten eingeben (größer 0).")
    cur = con.cursor()
    cur.execute(
        "INSERT INTO time_entries(task_id, manual_minutes, note, created_at) VALUES (?, ?, ?, ?)",
        (task_id, total_minutes, note, now_str()),
    )
    con.commit()


def get_task_summary(con):
    cur = con.cursor()
    cur.execute(
        """
        WITH base AS (
            SELECT t.id AS task_id,
                   t.name AS name,
                   t.planned_minutes AS planned,
                   IFNULL(SUM(CASE WHEN te.manual_minutes>0 THEN te.manual_minutes ELSE 0 END), 0) AS manual_sum,
                   IFNULL(SUM(CASE WHEN te.manual_minutes=0 AND te.end_time IS NOT NULL THEN 
                        CAST((strftime('%s', te.end_time) - strftime('%s', te.start_time)) / 60 AS INTEGER)
                   ELSE 0 END), 0) AS done_sum
            FROM tasks t
            LEFT JOIN time_entries te ON te.task_id = t.id
            GROUP BY t.id, t.name
        ),
        running AS (
            SELECT te.task_id,
                   CAST((strftime('%s', ?) - strftime('%s', te.start_time)) / 60 AS INTEGER) AS running_minutes
            FROM time_entries te
            WHERE te.end_time IS NULL AND te.manual_minutes=0
        )
        SELECT b.task_id,
               b.name,
               b.manual_sum + b.done_sum + IFNULL(r.running_minutes, 0) AS total_minutes,
               CASE WHEN r.task_id IS NOT NULL THEN 1 ELSE 0 END AS running,
               b.planned AS planned_minutes,
               (b.manual_sum + b.done_sum + IFNULL(r.running_minutes, 0)) - b.planned AS delta_minutes
        FROM base b
        LEFT JOIN running r ON r.task_id = b.task_id
        ORDER BY b.name COLLATE NOCASE
        """,
        (now_str(),),
    )
    return cur.fetchall()


def get_entries_for_task(con, task_id):
    cur = con.cursor()
    cur.execute(
        """
        SELECT id, start_time, end_time, manual_minutes, note, created_at
        FROM time_entries
        WHERE task_id = ?
        ORDER BY id DESC
        """,
        (task_id,),
    )
    return cur.fetchall()


# =========================
# Reporting
# =========================
def parse_date(dstr):
    return datetime.strptime(dstr, DATE_ONLY)


def clamp_range_minutes(start_str, end_str, manual_minutes, created_at_str, r_from, r_to):
    if manual_minutes and manual_minutes > 0:
        if created_at_str is None:
            return manual_minutes
        cat = datetime.strptime(created_at_str, DATE_FMT)
        return manual_minutes if r_from <= cat <= r_to else 0
    if start_str is None:
        return 0
    start = datetime.strptime(start_str, DATE_FMT)
    end = datetime.strptime(end_str, DATE_FMT) if end_str else datetime.now()
    s = max(start, r_from)
    e = min(end, r_to)
    if e <= s:
        return 0
    return int((e - s).total_seconds() // 60)


def build_report(con, date_from_str, date_to_str):
    r_from = parse_date(date_from_str)
    r_to = parse_date(date_to_str) + timedelta(days=1) - timedelta(seconds=1)
    cur = con.cursor()
    cur.execute(
        """
        SELECT t.id, t.name, t.planned_minutes,
               te.start_time, te.end_time, te.manual_minutes, te.note, te.created_at
        FROM tasks t
        LEFT JOIN time_entries te ON te.task_id = t.id
        ORDER BY t.name COLLATE NOCASE, te.id
        """
    )
    rows = cur.fetchall()
    per_task, details = {}, []
    for tid, name, planned, start, end, manual, note, cat in rows:
        mins = clamp_range_minutes(start, end, manual or 0, cat, r_from, r_to)
        if mins <= 0:
            continue
        per_task.setdefault((tid, name, planned or 0), 0)
        per_task[(tid, name, planned or 0)] += mins
        details.append([tid, name, start or "", end or "", int(manual or 0), note or "", cat or "", int(mins)])
    return per_task, details, r_from, r_to


# =========================
# GUI
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Zeiterfassung")
        self.geometry("1040x640")
        self.minsize(980, 600)
        self.con = init_db()

        left = ttk.Frame(self, padding=10); left.pack(side=tk.LEFT, fill=tk.BOTH, expand=False)
        ttk.Label(left, text="Tasks", font=("Arial", 12, "bold")).pack(anchor="w")

        self.task_tree = ttk.Treeview(
            left, columns=("id", "name", "sum", "planned", "delta", "state"),
            show="headings", height=18
        )
        for cid, label, w, anchor in [
            ("id","ID",0,tk.W), ("name","Name",260,tk.W),
            ("sum","Ist (h:m)",110,tk.CENTER), ("planned","Plan (h:m)",110,tk.CENTER),
            ("delta","Delta (h:m)",110,tk.CENTER), ("state","Status",100,tk.CENTER)
        ]:
            self.task_tree.heading(cid, text=label)
            self.task_tree.column(cid, width=w, anchor=anchor, stretch=(cid!="id"))
        self.task_tree.pack(fill=tk.BOTH, expand=True, pady=(6,8))
        self.task_tree.bind("<<TreeviewSelect>>", self.on_task_select)

        add_frame = ttk.LabelFrame(left, text="Neuen Task anlegen"); add_frame.pack(fill=tk.X)
        self.new_task_var = tk.StringVar(); self.new_pl_h = tk.StringVar(value="0"); self.new_pl_m = tk.StringVar(value="0")
        ttk.Label(add_frame, text="Name:").grid(row=0, column=0, padx=4, pady=4, sticky="e")
        ttk.Entry(add_frame, textvariable=self.new_task_var, width=22).grid(row=0, column=1, padx=4, pady=4)
        ttk.Label(add_frame, text="Plan h:").grid(row=0, column=2, padx=4, pady=4, sticky="e")
        ttk.Entry(add_frame, textvariable=self.new_pl_h, width=6).grid(row=0, column=3, padx=4, pady=4)
        ttk.Label(add_frame, text="Plan m:").grid(row=0, column=4, padx=4, pady=4, sticky="e")
        ttk.Entry(add_frame, textvariable=self.new_pl_m, width=6).grid(row=0, column=5, padx=4, pady=4)
        ttk.Button(add_frame, text="Task anlegen", command=self.ui_add_task).grid(row=0, column=6, padx=8, pady=4)

        btns = ttk.Frame(left); btns.pack(fill=tk.X, pady=(10,0))
        ttk.Button(btns, text="Start", command=self.ui_start).pack(side=tk.LEFT)
        ttk.Button(btns, text="Stop", command=self.ui_stop).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="Aktualisieren", command=self.refresh).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="Bericht...", command=self.ui_report_dialog).pack(side=tk.LEFT, padx=6)

        exp = ttk.Frame(left); exp.pack(fill=tk.X, pady=(10,0))
        ttk.Button(exp, text="Export Excel/CSV (alle)", command=self.ui_export_all).pack(side=tk.LEFT)
        ttk.Button(exp, text="Export HTML (alle)", command=self.ui_export_html_all).pack(side=tk.LEFT, padx=6)

        right = ttk.Frame(self, padding=10); right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(right, text="Details zum ausgewählten Task", font=("Arial", 12, "bold")).pack(anchor="w")

        manual = ttk.LabelFrame(right, text="Manuelle Zeit hinzufügen"); manual.pack(fill=tk.X, pady=(6,8))
        self.h_var = tk.StringVar(value="0"); self.m_var = tk.StringVar(value="0"); self.note_var = tk.StringVar(value="Manuell")
        ttk.Label(manual, text="Stunden:").grid(row=0, column=0, padx=4, pady=4, sticky="e")
        ttk.Entry(manual, width=6, textvariable=self.h_var).grid(row=0, column=1, padx=4, pady=4)
        ttk.Label(manual, text="Minuten:").grid(row=0, column=2, padx=4, pady=4, sticky="e")
        ttk.Entry(manual, width=6, textvariable=self.m_var).grid(row=0, column=3, padx=4, pady=4)
        ttk.Label(manual, text="Notiz:").grid(row=0, column=4, padx=4, pady=4, sticky="e")
        ttk.Entry(manual, width=20, textvariable=self.note_var).grid(row=0, column=5, padx=4, pady=4)
        ttk.Button(manual, text="Hinzufügen", command=self.ui_add_manual).grid(row=0, column=6, padx=8, pady=4)

        self.entries_tree = ttk.Treeview(
            right, columns=("start","end","manual","note","created"), show="headings"
        )
        for cid,label,w,anchor in [
            ("start","Start",150,tk.W), ("end","Ende",150,tk.W),
            ("manual","Manuell (min)",110,tk.CENTER), ("note","Notiz",180,tk.W),
            ("created","Erfasst am",150,tk.W)
        ]:
            self.entries_tree.heading(cid, text=label)
            self.entries_tree.column(cid, width=w, anchor=anchor)
        self.entries_tree.pack(fill=tk.BOTH, expand=True)

        self.status_var = tk.StringVar(value="Bereit.")
        ttk.Label(self, textvariable=self.status_var, anchor="w").pack(side=tk.BOTTOM, fill=tk.X)

        self.refresh()

    # ---- helpers ----
    def get_selected_task_id(self):
        sel = self.task_tree.selection()
        if not sel: return None
        values = self.task_tree.item(sel[0], "values")
        return int(values[0]) if values else None

    def refresh(self):
        for i in self.task_tree.get_children():
            self.task_tree.delete(i)
        for task_id, name, total_minutes, running, planned, delta in get_task_summary(self.con):
            h, m = divmod(int(total_minutes or 0), 60)
            ph, pm = divmod(int(planned or 0), 60)
            dh, dm = divmod(int(delta or 0), 60)
            state = "läuft" if running else "inaktiv"
            self.task_tree.insert("", tk.END, values=(task_id, name, f"{h:02d}:{m:02d}",
                                                      f"{ph:02d}:{pm:02d}", f"{dh:+03d}:{dm:02d}", state))
        self.load_entries(None)
        active = get_active_entry(self.con)
        self.status_var.set(f"Aktiv: Task-ID {active[1]} seit {active[2]}" if active else "Kein Task aktiv.")

    def on_task_select(self, _=None):
        self.load_entries(self.get_selected_task_id())

    def load_entries(self, task_id):
        for i in self.entries_tree.get_children():
            self.entries_tree.delete(i)
        if task_id is None: return
        for _id, start, end, manual, note, created_at in get_entries_for_task(self.con, task_id):
            self.entries_tree.insert("", tk.END, values=(start or "", end or "", int(manual or 0),
                                                         note or "", created_at or ""))

    # ---- UI Actions ----
    def ui_add_task(self):
        name = self.new_task_var.get().strip()
        if not name:
            messagebox.showwarning("Hinweis", "Bitte einen Task-Namen eingeben."); return
        try:
            ph = int(self.new_pl_h.get() or 0); pm = int(self.new_pl_m.get() or 0)
            create_task(self.con, name, ph, pm)
            self.new_task_var.set(""); self.new_pl_h.set("0"); self.new_pl_m.set("0")
            self.refresh()
        except sqlite3.IntegrityError:
            messagebox.showerror("Fehler", "Task existiert bereits.")
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    def ui_start(self):
        tid = self.get_selected_task_id()
        if tid is None: messagebox.showinfo("Hinweis", "Bitte zuerst einen Task auswählen."); return
        try:
            start_task(self.con, tid); self.refresh()
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    def ui_stop(self):
        tid = self.get_selected_task_id()
        if tid is None: messagebox.showinfo("Hinweis", "Bitte zuerst einen Task auswählen."); return
        try:
            stop_task(self.con, tid); self.refresh()
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    def ui_add_manual(self):
        tid = self.get_selected_task_id()
        if tid is None: messagebox.showinfo("Hinweis", "Bitte zuerst einen Task auswählen."); return
        try:
            h = int(self.h_var.get() or 0); m = int(self.m_var.get() or 0)
            note = self.note_var.get().strip() or "Manuell"
            add_manual_time(self.con, tid, h, m, note)
            self.h_var.set("0"); self.m_var.set("0")
            self.refresh(); self.load_entries(tid)
        except ValueError as e:
            messagebox.showwarning("Eingabe prüfen", str(e))
        except Exception as e:
            messagebox.showerror("Fehler", str(e))

    def ui_export_all(self):
        default_name = f"tasks_und_zeiten_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        target = filedialog.asksaveasfilename(
            title="Export speichern als", defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel-Datei (*.xlsx)", ".xlsx"), ("CSV-Datei (*.csv)", ".csv"), ("Alle Dateien", "*.*")]
        )
        if not target: return
        try:
            self.export_all_data(target)
            messagebox.showinfo("Export fertig", f"Export gespeichert:\n{target}")
        except Exception as e:
            messagebox.showerror("Export-Fehler", str(e))

    def ui_export_html_all(self):
        default_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        target = filedialog.asksaveasfilename(
            title="HTML-Report speichern als", defaultextension=".html",
            initialfile=default_name, filetypes=[("HTML-Datei (*.html)", ".html"), ("Alle Dateien", "*.*")]
        )
        if not target: return
        try:
            self.export_all_html(target)
            messagebox.showinfo("HTML-Report", f"Report gespeichert:\n{target}")
        except Exception as e:
            messagebox.showerror("Export-Fehler", str(e))

    def ui_report_dialog(self):
        dlg = tk.Toplevel(self); dlg.title("Bericht erstellen (Zeitraum)"); dlg.grab_set()
        ttk.Label(dlg, text="Von (YYYY-MM-DD):").grid(row=0, column=0, padx=6, pady=6, sticky="e")
        ttk.Label(dlg, text="Bis (YYYY-MM-DD):").grid(row=1, column=0, padx=6, pady=6, sticky="e")
        v_from = tk.StringVar(value=datetime.now().strftime(DATE_ONLY))
        v_to = tk.StringVar(value=datetime.now().strftime(DATE_ONLY))
        ttk.Entry(dlg, textvariable=v_from, width=14).grid(row=0, column=1, padx=6, pady=6)
        ttk.Entry(dlg, textvariable=v_to, width=14).grid(row=1, column=1, padx=6, pady=6)

        def do_excel():
            try:
                per_task, details, r_from, r_to = build_report(self.con, v_from.get(), v_to.get())
                default_name = f"bericht_{r_from.strftime('%Y%m%d')}_{r_to.strftime('%Y%m%d')}.xlsx"
                target = filedialog.asksaveasfilename(
                    title="Bericht speichern als (Excel/CSV)", defaultextension=".xlsx",
                    initialfile=default_name, filetypes=[("Excel-Datei (*.xlsx)", ".xlsx"), ("CSV-Datei (*.csv)", ".csv")]
                )
                if not target: return
                self.export_report(target, per_task, details, r_from, r_to)
                messagebox.showinfo("Bericht erstellt", f"Bericht gespeichert:\n{target}")
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("Fehler", str(e))

        def do_html():
            try:
                per_task, details, r_from, r_to = build_report(self.con, v_from.get(), v_to.get())
                default_name = f"bericht_{r_from.strftime('%Y%m%d')}_{r_to.strftime('%Y%m%d')}.html"
                target = filedialog.asksaveasfilename(
                    title="HTML-Report speichern als", defaultextension=".html",
                    initialfile=default_name, filetypes=[("HTML-Datei (*.html)", ".html"), ("Alle Dateien", "*.*")]
                )
                if not target: return
                self.export_report_html(target, per_task, details, r_from, r_to)
                messagebox.showinfo("HTML-Report erstellt", f"Report gespeichert:\n{target}")
                dlg.destroy()
            except Exception as e:
                messagebox.showerror("Fehler", str(e))

        row = ttk.Frame(dlg); row.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(row, text="Export Excel/CSV", command=do_excel).pack(side=tk.LEFT, padx=6)
        ttk.Button(row, text="Export HTML", command=do_html).pack(side=tk.LEFT, padx=6)

    # =========================
    # Exporte
    # =========================
    def export_all_data(self, target_path):
        cur = self.con.cursor()
        cur.execute(
            """
            SELECT t.id, t.name, t.planned_minutes,
                   te.start_time, te.end_time, te.manual_minutes, te.note
            FROM tasks t
            LEFT JOIN time_entries te ON te.task_id = t.id
            ORDER BY t.name COLLATE NOCASE, te.id
            """
        )
        rows = cur.fetchall()

        headers = ["Task-ID", "Task-Name", "Start", "Ende", "Manuell (min)", "Notiz", "Dauer (min)"]
        data = []
        for tid, name, planned, start, end, manual, note in rows:
            if manual and manual > 0:
                duration = int(manual)
            elif start and end:
                try:
                    duration = int((datetime.strptime(end, DATE_FMT) - datetime.strptime(start, DATE_FMT)).total_seconds() // 60)
                except Exception:
                    duration = 0
            else:
                duration = 0
            data.append([int(tid), name, start or "", end or "", int(manual or 0), note or "", int(duration)])

        # Summen (in Stunden, floats)
        summary_map = {}
        for tid, name, planned, *_ in rows:
            summary_map.setdefault((int(tid), name), {"planned_h": float((planned or 0)/60), "ist_h": 0.0})
        for tid, name, *_rest, duration in data:
            summary_map[(int(tid), name)]["ist_h"] += float(duration)/60.0
        for k in list(summary_map.keys()):
            summary_map[k]["planned_h"] = round(summary_map[k]["planned_h"], 2)
            summary_map[k]["ist_h"] = round(summary_map[k]["ist_h"], 2)
            summary_map[k]["delta_h"] = round(summary_map[k]["ist_h"] - summary_map[k]["planned_h"], 2)

        # Gesamtwerte
        total_plan = round(sum(v["planned_h"] for v in summary_map.values()), 2)
        total_ist = round(sum(v["ist_h"] for v in summary_map.values()), 2)
        total_delta = round(total_ist - total_plan, 2)

        ext = os.path.splitext(target_path)[1].lower()
        if ext == ".csv":
            import csv
            with open(target_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(headers); w.writerows(data); w.writerow([])
                w.writerow(["Summen pro Task (Stunden)"])
                w.writerow(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
                for (tid,name),v in summary_map.items():
                    w.writerow([tid,name,v["planned_h"],v["ist_h"],v["delta_h"]])
                # GESAMT
                w.writerow([])
                w.writerow(["GESAMT","","Plan (h)","Ist (h)","Delta (h)"])
                w.writerow(["","", total_plan, total_ist, total_delta])
            return

        # Excel
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference

            wb = Workbook()

            ws = wb.active; ws.title = "Einträge"
            ws.append(headers)
            for r in data: ws.append(r)
            for c in range(1, len(headers)+1): ws.column_dimensions[get_column_letter(c)].width = 18

            ws2 = wb.create_sheet("Summen")
            ws2.append(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
            items = sorted(summary_map.items(), key=lambda kv: kv[1]["ist_h"], reverse=True)
            for (tid, name), v in items:
                ws2.append([int(tid), name, float(v["planned_h"]), float(v["ist_h"]), float(v["delta_h"])])
            for c in range(1,6): ws2.column_dimensions[get_column_letter(c)].width = 20

            # GESAMT-Zeile
            if ws2.max_row >= 2:
                last = ws2.max_row
                ws2.append([])
                ws2.append(["GESAMT","",
                            f"=ROUND(SUM(C2:C{last}),2)",
                            f"=ROUND(SUM(D2:D{last}),2)",
                            f"=ROUND(D{last+2}-C{last+2},2)"])

            # robustes Diagramm (Standardoptionen)
            last = ws2.max_row
            # Diagramm nur über die Task-Zeilen (ohne GESAMT/Leerzeilen)
            last_tasks_row = None
            for r in range(ws2.max_row, 1, -1):
                if ws2.cell(row=r, column=1).value not in (None, "GESAMT"):
                    last_tasks_row = r
                    break
            if last_tasks_row and last_tasks_row >= 3:
                chart = BarChart()
                chart.title = "Plan vs. Ist (Stunden) pro Task"
                chart.y_axis.title = "Stunden"; chart.x_axis.title = "Task"
                data_ref = Reference(ws2, min_col=3, min_row=1, max_col=4, max_row=last_tasks_row)  # C..D (Plan, Ist)
                cats_ref = Reference(ws2, min_col=2, min_row=2, max_row=last_tasks_row)            # Task-Name
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws_chart = wb.create_sheet("Diagramm")
                ws_chart.add_chart(chart, "A1")

            wb.save(target_path)
        except Exception as e:
            import csv
            base, _ = os.path.splitext(target_path)
            csv_path = base + ".csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(headers); w.writerows(data); w.writerow([])
                w.writerow(["Summen pro Task (Stunden)"])
                w.writerow(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
                for (tid,name),v in summary_map.items():
                    w.writerow([tid,name,v["planned_h"],v["ist_h"],v["delta_h"]])
                w.writerow([])
                w.writerow(["GESAMT","","Plan (h)","Ist (h)","Delta (h)"])
                w.writerow(["","", total_plan, total_ist, total_delta])
            raise RuntimeError(f"Excel-Export fehlgeschlagen, CSV-Fallback gespeichert: {csv_path}\nGrund: {e}")

    def export_all_html(self, target_path):
        cur = self.con.cursor()
        cur.execute(
            """
            SELECT t.id, t.name, t.planned_minutes,
                   te.start_time, te.end_time, te.manual_minutes, te.note, te.created_at
            FROM tasks t
            LEFT JOIN time_entries te ON te.task_id = t.id
            ORDER BY t.name COLLATE NOCASE, te.id
            """
        )
        rows = cur.fetchall()

        tasks = {}
        for tid, name, planned_min, start, end, manual, note, created_at in rows:
            if tid not in tasks:
                tasks[tid] = {"id": int(tid), "name": name, "plan_h": round((planned_min or 0)/60, 2),
                              "ist_h": 0.0, "entries": []}
            if manual and manual > 0:
                minutes = int(manual)
            elif start and end:
                try:
                    minutes = int((datetime.strptime(end, DATE_FMT) - datetime.strptime(start, DATE_FMT)).total_seconds() // 60)
                except Exception:
                    minutes = 0
            else:
                minutes = 0
            if any([start, end, manual, note, created_at]):
                tasks[tid]["entries"].append({
                    "start": start or "", "end": end or "",
                    "manual_min": int(manual or 0), "note": note or "",
                    "created_at": created_at or "", "minutes": int(minutes)
                })
            tasks[tid]["ist_h"] += float(minutes)/60.0

        for t in tasks.values():
            t["ist_h"] = round(float(t["ist_h"]), 2); t["delta_h"] = round(t["ist_h"] - t["plan_h"], 2)

        ordered = sorted(tasks.values(), key=lambda x: x["ist_h"], reverse=True)
        total_plan_h = round(sum(t['plan_h'] for t in ordered), 2)
        total_ist_h = round(sum(t['ist_h'] for t in ordered), 2)
        payload = {
            "labels": [t["name"] for t in ordered],
            "plan":   [t["plan_h"] for t in ordered],
            "ist":    [t["ist_h"]  for t in ordered],
            "details": {t["name"]: t for t in ordered},
            "total_plan_h": total_plan_h,
            "total_ist_h": total_ist_h,
            "generated_at": now_str(),
        }
        data_json = json.dumps(payload).replace("</", "<\\/")
        html_text = build_html_page(data_json, title="Plan vs. Ist – Gesamt")
        with open(target_path, "w", encoding="utf-8") as f: f.write(html_text)

    def export_report(self, target_path, per_task, details, r_from, r_to):
        items_h = []
        for (tid, name, planned_min), ist_min in per_task.items():
            pl = round((planned_min or 0)/60, 2); ist = round((ist_min or 0)/60, 2)
            items_h.append((int(tid), name, float(pl), float(ist), float(round(ist - pl, 2))))

        # Totals für CSV/Excel
        total_plan = round(sum(x[2] for x in items_h), 2)
        total_ist  = round(sum(x[3] for x in items_h), 2)
        total_delta= round(total_ist - total_plan, 2)

        ext = os.path.splitext(target_path)[1].lower()
        if ext == ".csv":
            import csv
            with open(target_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(["Bericht von", r_from.strftime(DATE_FMT), "bis", r_to.strftime(DATE_FMT)])
                w.writerow([]); w.writerow(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
                for tid,name,pl,ist,dl in sorted(items_h, key=lambda x: x[3], reverse=True):
                    w.writerow([tid,name,pl,ist,dl])
                # GESAMT
                w.writerow([])
                w.writerow(["GESAMT","","Plan (h)","Ist (h)","Delta (h)"])
                w.writerow(["","", total_plan, total_ist, total_delta])
                w.writerow([]); w.writerow(["Details:"])
                w.writerow(["Task-ID","Task-Name","Start","Ende","Manuell (min)","Notiz","Erfasst am","Minuten in Zeitraum"])
                for row in details: w.writerow(row)
            return

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference

            wb = Workbook()
            ws = wb.active; ws.title = "Bericht"
            ws.append(["Berichtszeitraum", r_from.strftime(DATE_FMT), "bis", r_to.strftime(DATE_FMT)])
            ws.append([]); ws.append(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
            for tid,name,pl,ist,dl in sorted(items_h, key=lambda x: x[3], reverse=True):
                ws.append([int(tid), name, float(pl), float(ist), float(dl)])
            for c in range(1,6): ws.column_dimensions[get_column_letter(c)].width = 18

            # GESAMT-Zeile (unter der Tabelle)
            start_row = 4
            end_row = ws.max_row
            if end_row >= start_row:
                ws.append([])
                ws.append(["GESAMT","",
                           f"=ROUND(SUM(C{start_row}:C{end_row}),2)",
                           f"=ROUND(SUM(D{start_row}:D{end_row}),2)",
                           f"=ROUND(D{end_row+2}-C{end_row+2},2)"])

            last = ws.max_row
            if last >= 5:
                chart = BarChart()
                chart.title = "Plan vs. Ist (Stunden) – Bericht"
                chart.y_axis.title = "Stunden"; chart.x_axis.title = "Task"
                # Diagramm nur über die Task-Zeilen (ohne GESAMT/Leer)
                chart_max_row = end_row
                data_ref = Reference(ws, min_col=3, min_row=3, max_col=4, max_row=chart_max_row)
                cats_ref = Reference(ws, min_col=2, min_row=4, max_row=chart_max_row)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws_chart = wb.create_sheet("Diagramm Bericht")
                ws_chart.add_chart(chart, "A1")

            ws2 = wb.create_sheet("Details")
            ws2.append(["Task-ID","Task-Name","Start","Ende","Manuell (min)","Notiz","Erfasst am","Minuten in Zeitraum"])
            for row in details: ws2.append(row)
            for c in range(1,9): ws2.column_dimensions[get_column_letter(c)].width = 20

            wb.save(target_path)
        except Exception as e:
            import csv
            base,_ = os.path.splitext(target_path); csv_path = base + ".csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(["Bericht von", r_from.strftime(DATE_FMT), "bis", r_to.strftime(DATE_FMT)])
                w.writerow([]); w.writerow(["Task-ID","Task-Name","Plan (h)","Ist (h)","Delta (h)"])
                for tid,name,pl,ist,dl in sorted(items_h, key=lambda x: x[3], reverse=True):
                    w.writerow([tid,name,pl,ist,dl])
                w.writerow([])
                w.writerow(["GESAMT","","Plan (h)","Ist (h)","Delta (h)"])
                w.writerow(["","", total_plan, total_ist, total_delta])
                w.writerow([]); w.writerow(["Details:"])
                w.writerow(["Task-ID","Task-Name","Start","Ende","Manuell (min)","Notiz","Erfasst am","Minuten in Zeitraum"])
                for row in details: w.writerow(row)
            raise RuntimeError(f"Excel-Export fehlgeschlagen, CSV-Fallback gespeichert: {csv_path}\nGrund: {e}")

    def export_report_html(self, target_path, per_task, details, r_from, r_to):
        tasks = {}
        for (tid, name, planned_min), ist_min in per_task.items():
            tasks[name] = {"plan_h": round((planned_min or 0)/60, 2), "ist_h": round((ist_min or 0)/60, 2), "entries": []}
        for tid, name, start, end, manual_min, note, created_at, minutes in details:
            tasks.setdefault(name, {"plan_h": 0.0, "ist_h": 0.0, "entries": []})
            tasks[name]["entries"].append({
                "start": start, "end": end, "manual_min": int(manual_min or 0),
                "note": note, "created_at": created_at, "minutes": int(minutes or 0)
            })

        ordered = sorted(((n,v) for n,v in tasks.items()), key=lambda x: x[1]["ist_h"], reverse=True)
        labels = [n for n,_ in ordered]; plan = [v["plan_h"] for _,v in ordered]; ist = [v["ist_h"] for _,v in ordered]
        details_map = {n:v for n,v in ordered}

        total_plan_h = round(sum(v["plan_h"] for _, v in ordered), 2)
        total_ist_h  = round(sum(v["ist_h"]  for _, v in ordered), 2)

        payload = {
            "labels": labels, "plan": plan, "ist": ist, "details": details_map,
            "generated_at": now_str(), "range_from": r_from.strftime("%Y-%m-%d"), "range_to": r_to.strftime("%Y-%m-%d"),
            "total_plan_h": total_plan_h, "total_ist_h": total_ist_h
        }
        data_json = json.dumps(payload).replace("</", "<\\/")
        title = f"Zeitraum: {r_from.strftime('%Y-%m-%d')} bis {r_to.strftime('%Y-%m-%d')}"
        with open(target_path, "w", encoding="utf-8") as f:
            f.write(build_html_page(data_json, title=title, show_range=True))


# =========================
# HTML-Template (kompakt & Summen oben)
# =========================
def build_html_page(data_json, title="Plan vs. Ist", show_range=False):
    template = r"""<!doctype html>
<html lang="de"><head>
<meta charset="utf-8" /><meta name="viewport" content="width=device-width, initial-scale=1" />
<title>%%TITLE%%</title>
<link rel="preconnect" href="https://cdn.jsdelivr.net" /><link rel="dns-prefetch" href="https://cdn.jsdelivr.net" />
<style>
  :root{--bg:#0b0f19;--panel:#111827;--panel2:#0f172a;--text:#e5e7eb;--muted:#9ca3af;--border:#1f2937;--blue:#4472C4;--orange:#ED7D31;--green:#22c55e;--red:#ef4444}
  html,body{margin:0;background:var(--bg);color:var(--text);font:14px/1.4 system-ui,-apple-system,Segoe UI,Roboto,Ubuntu,Helvetica,Arial}
  header{padding:12px 16px;border-bottom:1px solid var(--border);background:var(--panel)}
  header h1{margin:0;font-size:18px} header .meta{color:var(--muted)}
  .wrap{max-width:1200px;margin:0 auto;padding:14px}
  .card{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:12px;margin-bottom:14px}
  .row{display:grid;grid-template-columns:2fr 1fr;gap:12px}
  input[type=search]{width:100%;background:var(--panel2);border:1px solid var(--border);color:var(--text);padding:8px 10px;border-radius:8px}
  table{width:100%;border-collapse:collapse} th,td{padding:6px 8px;border-bottom:1px solid var(--border)} th{color:var(--muted);text-align:left;font-weight:600}
  .muted{color:var(--muted)}
  canvas{background:transparent;height:420px}
  /* Summenkacheln */
  .stats{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
  .stat{background:var(--panel2);border:1px solid var(--border);border-radius:10px;padding:10px}
  .stat h4{margin:0 0 6px;font-size:12px;color:var(--muted);font-weight:600;text-transform:uppercase;letter-spacing:.04em}
  .stat .val{font-size:22px;font-variant-numeric:tabular-nums}
  .badge{display:inline-block;padding:2px 8px;border-radius:999px;border:1px solid var(--border);font-size:12px;margin-left:8px}
  .pos{color:var(--green)} .neg{color:var(--red)}
</style>
</head>
<body>
<header>
  <h1>Plan vs. Ist (Stunden)%%RANGETAG%%</h1>
  <div class="meta" id="meta"></div>
</header>

<div class="wrap">

  <!-- Summen oben -->
  <div class="card">
    <div class="stats">
      <div class="stat">
        <h4>Plan (gesamt)</h4>
        <div class="val" id="sumPlan">–</div>
      </div>
      <div class="stat">
        <h4>Ist (gesamt)</h4>
        <div class="val" id="sumIst">–</div>
      </div>
      <div class="stat">
        <h4>Delta (Ist − Plan)</h4>
        <div class="val" id="sumDelta">– <span id="sumDeltaBadge" class="badge"></span></div>
      </div>
    </div>
  </div>

  <div class="card">
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px;">
      <input id="filter" type="search" placeholder="Tasks filtern (live)…" />
      <div class="muted">Legende klicken = Serien ein/aus • Balken klicken = Details</div>
    </div>
    <canvas id="bar"></canvas>
  </div>

  <div class="row">
    <div class="card">
      <h3 style="margin:0 0 8px;">Details: <span id="selName" class="muted">– nichts ausgewählt –</span></h3>
      <div style="overflow:auto;max-height:50vh;">
        <table id="detailTable"><thead>
          <tr><th>Start</th><th>Ende</th><th>Manuell (min)</th><th>Notiz</th><th>Erfasst am</th><th>Minuten</th></tr>
        </thead><tbody></tbody></table>
      </div>
    </div>
    <div class="card">
      <h3 style="margin:0 0 8px;">Mini-Chart</h3>
      <canvas id="mini"></canvas>
    </div>
  </div>
</div>

<script>window.__PAYLOAD__=%%DATA_JSON%%;</script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js" crossorigin="anonymous"></script>
<script>
(function(){
  const p=window.__PAYLOAD__, L=p.labels? p.labels.slice():[], PL=p.plan? p.plan.slice():[], IS=p.ist? p.ist.slice():[], D=p.details||{};
  const meta=document.getElementById('meta');
  meta.textContent=(p.range_from&&p.range_to)?`Zeitraum: ${p.range_from} bis ${p.range_to} • generiert: ${p.generated_at}`:`generiert: ${p.generated_at}`;

  // Summen befüllen
  const totalPlan=(typeof p.total_plan_h==='number')? p.total_plan_h : (PL.reduce((a,b)=>a+(+b||0),0));
  const totalIst =(typeof p.total_ist_h ==='number')? p.total_ist_h  : (IS.reduce((a,b)=>a+(+b||0),0));
  const delta=+(totalIst - totalPlan).toFixed(2);
  setTxt('sumPlan',  totalPlan.toFixed(2)+' h');
  setTxt('sumIst',   totalIst.toFixed(2)+' h');
  setTxt('sumDelta', (delta>=0? '+' : '') + delta.toFixed(2)+' h');
  const badge=document.getElementById('sumDeltaBadge');
  if (badge){
    badge.textContent = delta>=0 ? 'über Plan' : 'unter Plan';
    badge.className = 'badge ' + (delta>=0? 'pos':'neg');
  }

  const BLUE=get('--blue'), ORANGE=get('--orange');
  const chart=new Chart(document.getElementById('bar').getContext('2d'),{
    type:'bar',
    data:{labels:L,datasets:[
      {label:'Plan (h)',data:PL,backgroundColor:BLUE},
      {label:'Ist (h)', data:IS, backgroundColor:ORANGE}
    ]},
    options:{
      responsive:true, maintainAspectRatio:true,
      plugins:{ legend:{ position:'right', labels:{color:'#e5e7eb'}},
                tooltip:{callbacks:{label:c=>`${c.dataset.label}: ${c.formattedValue} h`}} },
      scales:{ x:{ticks:{color:'#9ca3af',maxRotation:45,autoSkip:true,autoSkipPadding:8},grid:{display:false}},
               y:{ticks:{color:'#9ca3af'},grid:{color:'rgba(255,255,255,.06)'},title:{display:true,text:'Stunden',color:'#9ca3af'}} },
      onClick:(e,els)=>{ if(!els.length) return; select(chart.data.labels[els[0].index]); }
    }
  });

  document.getElementById('filter').addEventListener('input',()=>{
    const q=event.target.value.trim().toLowerCase(); const l=[],pl=[],is=[];
    for(let i=0;i<L.length;i++){const n=L[i]; if(!q||n.toLowerCase().includes(q)){l.push(n);pl.push(PL[i]);is.push(IS[i]);}}
    chart.data.labels=l; chart.data.datasets[0].data=pl; chart.data.datasets[1].data=is; chart.update();
  });

  const tbody=document.querySelector('#detailTable tbody'), sel=document.getElementById('selName');
  const mini=new Chart(document.getElementById('mini').getContext('2d'),{
    type:'bar', data:{labels:['Plan','Ist'],datasets:[{label:'Stunden',data:[0,0],backgroundColor:[BLUE,ORANGE]}]},
    options:{responsive:true,maintainAspectRatio:true,plugins:{legend:{display:false}}}
  });

  function select(name){
    sel.textContent=name; const t=D[name]; if(!t){tbody.innerHTML=''; mini.data.datasets[0].data=[0,0]; mini.update(); return;}
    tbody.innerHTML=''; for(const e of (t.entries||[])){
      const tr=document.createElement('tr');
      tr.innerHTML=`<td>${e.start??''}</td><td>${e.end??''}</td><td>${e.manual_min??0}</td><td>${e.note??''}</td><td>${e.created_at??''}</td><td>${e.minutes??0}</td>`;
      tbody.appendChild(tr);
    }
    mini.data.datasets[0].data=[t.plan_h||0,t.ist_h||0]; mini.update();
  }
  function get(v){return getComputedStyle(document.documentElement).getPropertyValue(v).trim()||'#888';}
  function setTxt(id, txt){const el=document.getElementById(id); if(el) el.firstChild ? (el.firstChild.nodeValue=txt) : (el.textContent=txt);}
})();
</script>
</body></html>
"""
    return (template
            .replace("%%TITLE%%", html.escape(title))
            .replace("%%RANGETAG%%", " – Zeitraum" if show_range else "")
            .replace("%%DATA_JSON%%", data_json))


if __name__ == "__main__":
    app = App()
    app.mainloop()
